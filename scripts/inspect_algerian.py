"""Extract Algerian feed data from the Excel file."""
import openpyxl
import json
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook("/home/z/my-project/upload/Table_de_composition_et_de_la_valeur_nutritive_des_matieres_premiere.xlsx", data_only=True)

# Examine first 2 sheets in detail
for sheet_name in wb.sheetnames[:3]:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name} (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'='*80}")
    for r in range(1, min(ws.max_row + 1, 15)):
        row_str = []
        for c in range(1, min(ws.max_column + 1, 21)):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                col_letter = openpyxl.utils.get_column_letter(c)
                row_str.append(f"[{col_letter}{r}]={val}")
        if row_str:
            print(f"R{r}: " + " | ".join(row_str))
