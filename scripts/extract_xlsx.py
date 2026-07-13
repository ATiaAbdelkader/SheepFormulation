"""Extract data from AlimOVINS Excel file to understand structure."""
import openpyxl
import json
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook("/home/z/my-project/upload/AlimOVINSv5.1.xlsx", data_only=False)
wb_v = openpyxl.load_workbook("/home/z/my-project/upload/AlimOVINSv5.1.xlsx", data_only=True)

output = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws_v = wb_v[sheet_name]
    rows = []
    max_rows = min(ws.max_row, 60)  # cap to first 60 rows for inspection
    max_cols = min(ws.max_column, 30)
    for r in range(1, max_rows + 1):
        row_data = []
        for c in range(1, max_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell_v = ws_v.cell(row=r, column=c)
            val = cell.value
            val_v = cell_v.value
            if val is not None or val_v is not None:
                row_data.append({
                    "cell": cell.coordinate,
                    "formula": val if isinstance(val, str) and str(val).startswith("=") else None,
                    "value": val_v if val_v != val else (val if not (isinstance(val, str) and str(val).startswith("=")) else None),
                    "raw": val if not (isinstance(val, str) and str(val).startswith("=")) else None
                })
            else:
                row_data.append(None)
        rows.append(row_data)
    output[sheet_name] = {
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "rows_preview": rows
    }

# Save to file
with open("/home/z/my-project/scripts/xlsx_extract.json", "w") as f:
    json.dump(output, f, indent=2, default=str, ensure_ascii=False)

print(f"Extracted {len(output)} sheets")
print(f"Total size: {sum(len(json.dumps(v, default=str)) for v in output)} chars")
