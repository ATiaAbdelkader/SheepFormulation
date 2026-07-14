"""Extract all Algerian feeds from the Excel file and convert to OvinFormulation format.
Structure: Each sheet has blocks of 4 feeds (columns B-E, G-J, L-O, Q-T).
Each block has ~25 rows: header + composition + nutritive values.
"""
import openpyxl
import json
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook("/home/z/my-project/upload/Table_de_composition_et_de_la_valeur_nutritive_des_matieres_premiere.xlsx", data_only=True)

# Map sheet names to feed categories
SHEET_CATEGORIES = {
    "Graminees": "Céréales graminées",
    "Sous Produit des Bles Durs": "Sous-produits blés durs",
    "Sous Produit des Bles Tendre": "Sous-produits blés tendre",
    "Oleagineux": "Oléagineux",
    "Sous Produits d'Oleagineux :": "Sous-produits oléagineux",
    "Touteaux": "Tourteaux",
    "Sous Produit Phoenicicole": "Sous-produits phoenicicole (palmier)",
    "Proteagineux": "Protéagineux",
    "Produits forestiers": "Produits forestiers",
    "Farine Animal": "Farines animales",
    "Sous Produits Agroindustriels": "Sous-produits agroindustriels",
}

# Column groups: each group is (name_col, unit_col, moy_col, et_col)
# B=2, C=3, D=4, E=5 | G=7, H=8, I=9, J=10 | L=12, M=13, N=14, O=15 | Q=17, R=18, S=19, T=20
COLUMN_GROUPS = [(2, 3, 4, 5), (7, 8, 9, 10), (12, 13, 14, 15), (17, 18, 19, 20)]

# Row labels to extract (parameter name -> normalized key)
ROW_MAP = {
    "Matiere seche": "ms_pct",
    "Proteine Brutes": "protein_pct",
    "Cellulse Brute": "cb_pct",
    "Cellulose Brute": "cb_pct",
    "Matiere Grasse Brutes": "fat_pct",
    "Cendres Brutes": "ash_pct",
    "NDF": "ndf_pct",
    "ADF": "adf_pct",
    "ADL": "adl_pct",
    "Calicium": "ca_pct",
    "Phosphore": "p_pct",
    "Energie Brute": "energy_kcal",
    "UFL": "ufl",
    "UFV": "ufv",
    "PDIA": "pdia",
    "PDIN": "pdin",
    "PDIE": "pdie",
    "DMD": "dmd_pct",
}

def extract_feed_from_block(ws, start_row, name_col, moy_col):
    """Extract a single feed from a block starting at start_row."""
    # Feed name is at start_row, name_col+1 (e.g., C2 for first feed)
    name = ws.cell(row=start_row, column=name_col + 1).value
    if not name or str(name).strip() == "":
        return None
    name = str(name).strip()

    feed_data = {"name": name, "raw": {}}

    # Scan rows from start_row+3 (after header rows) to find composition
    for r in range(start_row + 3, min(start_row + 26, ws.max_row + 1)):
        label = ws.cell(row=r, column=name_col).value
        if not label:
            continue
        label_str = str(label).strip()

        # Check if this row matches any known parameter
        for key, param_key in ROW_MAP.items():
            if key.lower() in label_str.lower():
                moy_val = ws.cell(row=r, column=moy_col).value
                if moy_val is not None and isinstance(moy_val, (int, float)):
                    feed_data["raw"][param_key] = moy_val
                break

        # If we hit the next "Aliment" header, stop
        if "Aliment" in label_str and r > start_row + 3:
            break

    return feed_data

def convert_to_ovinformulation(feed, category):
    """Convert raw extracted data to OvinFormulation feed format."""
    raw = feed["raw"]

    # Determine kind based on category
    kind = "concentre"  # default
    if any(x in category.lower() for x in ["graminees", "fourrage", "paturage", "foin", "ensilage"]):
        kind = "fourrage"
    elif any(x in category.lower() for x in ["farine animal", "animal"]):
        kind = "concentre"

    # Convert Ca and P from % to g/kg, then estimate absorbable
    # Ca: % × 10 = g/kg; Caabs ≈ Ca × 0.30 (forage) or 0.40 (concentrate)
    ca_pct = raw.get("ca_pct")
    p_pct = raw.get("p_pct")

    ca_abs_coeff = 0.30 if kind == "fourrage" else 0.40
    p_abs_coeff = 0.65

    caabs = round(ca_pct * 10 * ca_abs_coeff, 2) if ca_pct is not None else None
    pabs = round(p_pct * 10 * p_abs_coeff, 2) if p_pct is not None else None

    # Estimate UEM (forages only, based on NDF/CB)
    uem = None
    if kind == "fourrage":
        cb = raw.get("cb_pct", 0)
        uem = round(max(0.5, min(1.5, 1.4 - 0.015 * cb)), 2) if cb else None

    return {
        "name": feed["name"],
        "kind": kind,
        "category": category,
        "ms_pct": raw.get("ms_pct"),
        "ufl": raw.get("ufl"),
        "ufv": raw.get("ufv"),
        "pdin": raw.get("pdin"),
        "pdie": raw.get("pdie"),
        "pdia": raw.get("pdia"),
        "protein_pct": raw.get("protein_pct"),
        "cb_pct": raw.get("cb_pct"),
        "fat_pct": raw.get("fat_pct"),
        "ash_pct": raw.get("ash_pct"),
        "ndf_pct": raw.get("ndf_pct"),
        "adf_pct": raw.get("adf_pct"),
        "adl_pct": raw.get("adl_pct"),
        "ca_pct": ca_pct,
        "p_pct": p_pct,
        "caabs": caabs,
        "pabs": pabs,
        "uem": uem,
        "energy_kcal": raw.get("energy_kcal"),
        "dmd_pct": raw.get("dmd_pct"),
        "price": None,  # No price data in Algerian table
        "source": "Table de composition algérienne (ITGC/INRAA)",
    }

# Extract all feeds
all_feeds = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    category = SHEET_CATEGORIES.get(sheet_name, sheet_name)
    print(f"\n--- Processing sheet: {sheet_name} ({category}) ---")

    # Some sheets have multiple blocks (e.g., Graminees has 2 blocks of 4 feeds)
    # Block 1 starts at row 2, Block 2 starts at row 27 (approx)
    # Check for multiple "Aliment" header rows
    block_starts = []
    for r in range(1, min(ws.max_row + 1, 80)):
        val = ws.cell(row=r, column=2).value  # Column B
        if val and "Aliment" in str(val):
            block_starts.append(r)

    print(f"  Found {len(block_starts)} block(s) at rows: {block_starts}")

    for block_start in block_starts:
        for group_idx, (name_col, unit_col, moy_col, et_col) in enumerate(COLUMN_GROUPS):
            # Check if this column group has data (feed name exists)
            feed_name_cell = ws.cell(row=block_start, column=name_col + 1).value
            if not feed_name_cell:
                continue

            feed = extract_feed_from_block(ws, block_start, name_col, moy_col)
            if feed and feed["raw"]:
                converted = convert_to_ovinformulation(feed, category)
                if converted["ufl"] is not None or converted["pdin"] is not None:
                    all_feeds.append(converted)
                    print(f"  ✓ {converted['name']}: UFL={converted['ufl']}, PDIN={converted['pdin']}, MS={converted['ms_pct']}%")

print(f"\n{'='*60}")
print(f"Total feeds extracted: {len(all_feeds)}")
print(f"By category:")
from collections import Counter
cat_counts = Counter(f["category"] for f in all_feeds)
for cat, count in cat_counts.items():
    print(f"  {cat}: {count}")

# Save to JSON
output_path = "/home/z/my-project/src/lib/algerian-feeds.json"
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(all_feeds, f, indent=2, ensure_ascii=False)
print(f"\nSaved to: {output_path}")
