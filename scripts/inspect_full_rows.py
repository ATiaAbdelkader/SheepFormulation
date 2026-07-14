"""Extract all Algerian feeds from the Excel file and convert to OvinFormulation format."""
import openpyxl
import json
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook("/home/z/my-project/upload/Table_de_composition_et_de_la_valeur_nutritive_des_matieres_premiere.xlsx", data_only=True)

# First, let's see ALL rows of one sheet to understand the full row list
ws = wb["Graminees"]
print("Full row list for Graminees sheet:")
for r in range(1, ws.max_row + 1):
    row_str = []
    for c in range(1, 6):  # Just first feed (columns B, C, D)
        val = ws.cell(row=r, column=c).value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            row_str.append(f"[{col_letter}]={val}")
    if row_str:
        print(f"R{r}: " + " | ".join(row_str))
