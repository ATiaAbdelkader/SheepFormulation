"""Extract complete AGNEAU (lamb) data and prepare final data.json for the web app."""
import openpyxl
import json
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook("/home/z/my-project/upload/AlimOVINSv5.1.xlsx", data_only=True)

# ============= AGNEAU LAMB REQUIREMENTS =============
# Rows 44-211, columns: C=Poids, D=GMQ, E=Potentiel, F=Sexe, G=Description, H=UFV, I=PDI, J=Pabs, K=Caabs
ws = wb["AGNEAU"]
agneaux = []
for r in range(44, 212):
    poids = ws.cell(row=r, column=3).value  # C
    gmq = ws.cell(row=r, column=4).value  # D
    potentiel = ws.cell(row=r, column=5).value  # E
    sexe = ws.cell(row=r, column=6).value  # F
    description = ws.cell(row=r, column=7).value  # G
    ufv = ws.cell(row=r, column=8).value  # H
    pdi = ws.cell(row=r, column=9).value  # I
    pabs = ws.cell(row=r, column=10).value  # J
    caabs = ws.cell(row=r, column=11).value  # K
    if poids and gmq:
        agneaux.append({
            "poids": str(poids).strip(),
            "gmq": str(gmq).strip(),
            "potentiel": str(potentiel).strip() if potentiel else "",
            "sexe": str(sexe).strip() if sexe else "",
            "description": str(description).strip() if description else "",
            "UFV": ufv if ufv != "ND" else None,
            "PDI": pdi if pdi != "ND" else None,
            "Pabs": pabs if pabs != "ND" else None,
            "Caabs": caabs if caabs != "ND" else None,
        })

print(f"Agneaux extracted: {len(agneaux)}")
print("Sample:", agneaux[6] if len(agneaux) > 6 else "none")

# ============= AGNEAU CONCENTRATE FEEDS =============
# Columns W-AE (23-31) for lamb feed composition
# W=Catégorie, X=%, Y=UFV, Z=PDIN, AA=PDIE, AB=Pabs, AC=Caabs, AD=MAT, AE=CB
agneau_concentres = []
for r in range(44, 90):
    cat = ws.cell(row=r, column=23).value  # W
    name = ws.cell(row=r, column=23).value  # W (name appears here)
    if not name or str(name).strip() == "":
        # Try alternative
        continue
    pct = ws.cell(row=r, column=24).value  # X
    ufv = ws.cell(row=r, column=25).value  # Y
    pdin = ws.cell(row=r, column=26).value  # Z
    pdie = ws.cell(row=r, column=27).value  # AA
    pabs = ws.cell(row=r, column=28).value  # AB
    caabs = ws.cell(row=r, column=29).value  # AC
    mat = ws.cell(row=r, column=30).value  # AD
    cb = ws.cell(row=r, column=31).value  # AE
    agneau_concentres.append({
        "name": str(name).strip(),
        "pct": pct,
        "ufv": ufv,
        "pdin": pdin,
        "pdie": pdie,
        "pabs": pabs,
        "caabs": caabs,
        "mat": mat,
        "cb": cb,
    })
# Deduplicate by name
seen = set()
agneau_concentres_unique = []
for c in agneau_concentres:
    if c["name"] not in seen:
        seen.add(c["name"])
        agneau_concentres_unique.append(c)
print(f"Agneau concentres unique: {len(agneau_concentres_unique)}")

# ============= PATURAGE CATEGORIES =============
ws = wb["PATURAGE"]
paturage_cats = []
for r in range(23, 27):
    cat = ws.cell(row=r, column=3).value  # C
    if cat:
        paturage_cats.append(str(cat).strip())
print(f"Paturage categories: {paturage_cats}")

paturage_density_options = []
for r in range(26, 30):
    val = ws.cell(row=r, column=8).value  # H
    if val:
        paturage_density_options.append(str(val).strip())
print(f"Paturage density options: {paturage_density_options}")

# ============= LOAD PREVIOUSLY EXTRACTED DATA =============
with open("/home/z/my-project/scripts/data.json", "r") as f:
    data = json.load(f)

# Add agneaux data
data["agneaux"] = agneaux
data["agneau_concentres"] = agneau_concentres_unique
data["paturage_categories"] = paturage_cats
data["paturage_density_options"] = paturage_density_options

# Save
with open("/home/z/my-project/scripts/data.json", "w") as f:
    json.dump(data, f, indent=2, ensure_ascii=False, default=str)

print(f"\n=== FINAL DATA SUMMARY ===")
print(f"Animals: {len(data['animals'])}")
print(f"Fourrages: {len(data['fourrages'])}")
print(f"Concentres: {len(data['concentres'])}")
print(f"CMVs: {len(data['cmvs'])}")
print(f"Besoins categories: {len(data['besoins_categories'])}")
print(f"Stock categories: {len(data['stock_categories'])}")
print(f"Agneaux: {len(data['agneaux'])}")
print(f"Agneau concentres: {len(data['agneau_concentres'])}")
print(f"Paturage categories: {len(data['paturage_categories'])}")
print(f"Paturage density: {len(data['paturage_density_options'])}")

# Print all unique animal types
animal_types = set(a["animal_type"] for a in data["animals"])
print(f"\nAnimal types: {sorted(animal_types)}")

# Print all unique stages
stages = set(a["stage"] for a in data["animals"] if a["stage"])
print(f"Stages: {sorted(stages)}")

# Print all unique weights
weights = set(a["weight"] for a in data["animals"] if a["weight"])
print(f"Weights: {sorted(weights)}")
