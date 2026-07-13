"""Extract data from key sheets to understand the full structure."""
import openpyxl
import json
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook("/home/z/my-project/upload/AlimOVINSv5.1.xlsx", data_only=True)

# Focus on key sheets first
key_sheets = ["ACCUEIL", "ANIMAL", "FOURRAGE", "CONCENTRE", "CMV", "RATION", "BILAN", "BESOINS", "MELANGE", "COUT", "AGNEAU", "PATURAGE", "NOUVEAU", "STOCKS", "PREVISION"]

for sheet_name in key_sheets:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name} (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'='*80}")
    max_rows = min(ws.max_row, 80)
    max_cols = min(ws.max_column, 30)
    for r in range(1, max_rows + 1):
        row_str = []
        for c in range(1, max_cols + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                row_str.append(f"[{openpyxl.utils.get_column_letter(c)}{r}]={val}")
        if row_str:
            print(f"R{r}: " + " | ".join(row_str))
