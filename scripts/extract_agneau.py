"""Extract AGNEAU (lamb) data and PATURAGE categories."""
import openpyxl
import json
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook("/home/z/my-project/upload/AlimOVINSv5.1.xlsx", data_only=True)

# AGNEAU - need to find the lamb requirements table
ws = wb["AGNEAU"]
print("="*80)
print("AGNEAU sheet - looking for lamb requirements table")
print("="*80)
for r in range(1, min(ws.max_row + 1, 225)):
    row_str = []
    for c in range(1, min(ws.max_column + 1, 32)):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            row_str.append(f"[{openpyxl.utils.get_column_letter(c)}{r}]={val}")
    if row_str:
        print(" | ".join(row_str))
