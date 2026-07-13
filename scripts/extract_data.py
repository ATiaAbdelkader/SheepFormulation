"""Extract structured data from AlimOVINS Excel for the web app."""
import openpyxl
import json
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.load_workbook("/home/z/my-project/upload/AlimOVINSv5.1.xlsx", data_only=True)

# ============= ANIMAL DATABASE =============
# Located in ANIMAL sheet, starting at row 39, columns C-I (Catégorie, UEM, UFL, PDI, Pabs, Caabs)
# Looking at the data, the format is:
# C: Category string (e.g., "Agnelle  30 kg   Croissance")
# D: number (some index?)
# E: UEM
# F: UFL
# G: PDI
# H: Pabs
# I: Caabs
ws = wb["ANIMAL"]
animals = []
for r in range(39, ws.max_row + 1):
    cat = ws.cell(row=r, column=3).value  # C
    if not cat or cat == "ND":
        continue
    uem = ws.cell(row=r, column=5).value  # E
    ufl = ws.cell(row=r, column=6).value  # F
    pdi = ws.cell(row=r, column=7).value  # G
    pabs = ws.cell(row=r, column=8).value  # H
    caabs = ws.cell(row=r, column=9).value  # I
    # Parse category string into type, weight, stage
    cat_str = str(cat).strip()
    # Examples: "Agnelle  30 kg   Croissance", "Brebis  70 kg   Gestante    Double      Semaines - 3 à - 4"
    parts = [p.strip() for p in cat_str.split("  ") if p.strip()]
    # Try to extract type, weight, stage, substage
    animal_type = parts[0] if parts else ""
    weight = ""
    stage = ""
    sub_stage = ""
    for p in parts[1:]:
        if "kg" in p and not weight:
            weight = p
        elif p in ["Croissance", "Entretien", "Flushing", "Gestante", "Allaitante", "Vide", "Engrais", "Traite", "Lactation"]:
            stage = p
        else:
            sub_stage = (sub_stage + " " + p).strip() if sub_stage else p
    animals.append({
        "category": cat_str,
        "animal_type": animal_type,
        "weight": weight,
        "stage": stage,
        "sub_stage": sub_stage,
        "UEM": uem,
        "UFL": ufl,
        "PDI": pdi,
        "Pabs": pabs,
        "Caabs": caabs,
    })

print(f"Animals extracted: {len(animals)}")
print("Sample:", animals[0] if animals else "none")

# ============= FOURRAGE DATABASE =============
# Located in FOURRAGE sheet starting row 78, columns C-M (Betterave fourragère example)
# C: Name, D: % MS, E: UEM, F: UEB, G: UFL, H: UFV, I: PDIN, J: PDIE, K: Pabs, L: Caabs, M: price?
ws = wb["FOURRAGE"]
fourrages = []
for r in range(78, ws.max_row + 1):
    name = ws.cell(row=r, column=3).value  # C
    if not name or str(name).strip() == "":
        continue
    fourrages.append({
        "name": str(name).strip(),
        "ms_pct": ws.cell(row=r, column=4).value,  # D % MS
        "uem": ws.cell(row=r, column=5).value,  # E
        "ueb": ws.cell(row=r, column=6).value,  # F
        "ufl": ws.cell(row=r, column=7).value,  # G
        "ufv": ws.cell(row=r, column=8).value,  # H
        "pdin": ws.cell(row=r, column=9).value,  # I
        "pdie": ws.cell(row=r, column=10).value,  # J
        "pabs": ws.cell(row=r, column=11).value,  # K
        "caabs": ws.cell(row=r, column=12).value,  # L
        "price": ws.cell(row=r, column=13).value,  # M
    })
print(f"Fourrages extracted: {len(fourrages)}")
print("Sample:", fourrages[0] if fourrages else "none")

# ============= CONCENTRE DATABASE =============
# Located in CONCENTRE sheet starting row 42, columns C-N
ws = wb["CONCENTRE"]
concentres = []
for r in range(42, ws.max_row + 1):
    name = ws.cell(row=r, column=3).value  # C
    if not name or str(name).strip() == "":
        continue
    concens = {
        "name": str(name).strip(),
        "ms_pct": ws.cell(row=r, column=4).value,  # D
        "ufl": ws.cell(row=r, column=5).value,  # E
        "pdin": ws.cell(row=r, column=6).value,  # F
        "pdie": ws.cell(row=r, column=7).value,  # G
        "pabs": ws.cell(row=r, column=8).value,  # H
        "caabs": ws.cell(row=r, column=9).value,  # I
        "price": ws.cell(row=r, column=11).value,  # K
        "price2": ws.cell(row=r, column=14).value,  # N
    }
    concens["price"] = concens["price"] if concens["price"] is not None else concens["price2"]
    del concens["price2"]
    concentres.append(concens)
print(f"Concentres extracted: {len(concentres)}")
print("Sample:", concentres[0] if concentres else "none")

# ============= CMV DATABASE =============
# Located in CMV sheet starting row 20, columns C-J (Type, F=Ca/P, G=Pabs, H=Caabs, I=P, J=Ca)
ws = wb["CMV"]
cmvs = []
for r in range(20, ws.max_row + 1):
    name = ws.cell(row=r, column=3).value  # C
    if not name or str(name).strip() == "":
        continue
    cmvs.append({
        "name": str(name).strip(),
        "ca_p_ratio": ws.cell(row=r, column=6).value,  # F
        "pabs_per_kg": ws.cell(row=r, column=7).value,  # G
        "caabs_per_kg": ws.cell(row=r, column=8).value,  # H
        "p_pct": ws.cell(row=r, column=9).value,  # I
        "ca_pct": ws.cell(row=r, column=10).value,  # J
    })
print(f"CMVs extracted: {len(cmvs)}")
print("Sample:", cmvs[0] if cmvs else "none")

# ============= AGNEAU DATABASE =============
ws = wb["AGNEAU"]
agneaux = []
# Look for the lamb requirements table
print("\nAGNEAU sheet structure:")
for r in range(1, min(ws.max_row + 1, 30)):
    row_str = []
    for c in range(1, min(ws.max_column + 1, 20)):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            row_str.append(f"[{openpyxl.utils.get_column_letter(c)}]={val}")
    if row_str:
        print(f"R{r}: " + " | ".join(row_str))

# ============= BESOINS DATA =============
# Lot categories and UGB equivalents
ws = wb["BESOINS"]
besoins_categories = []
for r in range(41, 52):
    cat = ws.cell(row=r, column=4).value  # D
    ugb = ws.cell(row=r, column=5).value  # E
    cons = ws.cell(row=r, column=6).value  # F
    if cat:
        besoins_categories.append({
            "category": cat,
            "ugb": ugb,
            "consumption_kg_ms_day": cons
        })

# Stock categories
stock_categories = []
for r in range(43, 50):
    cat = ws.cell(row=r, column=19).value  # S
    cons = ws.cell(row=r, column=20).value  # T
    if cat:
        stock_categories.append({
            "category": cat,
            "consumption_kg_ms_day": cons
        })

# ============= PATURAGE =============
ws = wb["PATURAGE"]
print("\nPATURAGE sheet structure (first 30 rows):")
for r in range(1, min(ws.max_row + 1, 30)):
    row_str = []
    for c in range(1, min(ws.max_column + 1, 20)):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            row_str.append(f"[{openpyxl.utils.get_column_letter(c)}]={val}")
    if row_str:
        print(f"R{r}: " + " | ".join(row_str))

# Save all data
output = {
    "animals": animals,
    "fourrages": fourrages,
    "concentres": concentres,
    "cmvs": cmvs,
    "besoins_categories": besoins_categories,
    "stock_categories": stock_categories,
}

with open("/home/z/my-project/scripts/data.json", "w") as f:
    json.dump(output, f, indent=2, ensure_ascii=False, default=str)

print(f"\nTotal animals: {len(animals)}")
print(f"Total fourrages: {len(fourrages)}")
print(f"Total concentres: {len(concentres)}")
print(f"Total cmvs: {len(cmvs)}")
print(f"Total besoins_categories: {len(besoins_categories)}")
print(f"Total stock_categories: {len(stock_categories)}")
print("Data saved to /home/z/my-project/scripts/data.json")
